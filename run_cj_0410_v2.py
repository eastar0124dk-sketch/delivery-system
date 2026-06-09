import sys
import win32com.client as win32
import openpyxl

src = r"C:\Users\dkchoi\Downloads\0410 dn add .xlsx"
dst = r"C:\Users\dkchoi\Downloads\메틀러토레도_CJ택배_매크로 (1).xlsm"

excel = win32.Dispatch("Excel.Application")
excel.Visible = True   # Visible so we can see if there's a dialog
excel.DisplayAlerts = False
excel.AskToUpdateLinks = False

macro_success = False
macro_error_msg = ""
last_row = 0
last_col = 0

try:
    print("Opening source file...", flush=True)
    wb_src = excel.Workbooks.Open(src, UpdateLinks=0, ReadOnly=True)
    print(f"Source opened. Sheet count: {wb_src.Sheets.Count}", flush=True)

    # List all sheets
    for i in range(1, wb_src.Sheets.Count + 1):
        print(f"  Sheet {i}: '{wb_src.Sheets(i).Name}'", flush=True)

    # Find "13시 add" sheet
    ws_src = None
    for i in range(1, wb_src.Sheets.Count + 1):
        sheet_name = wb_src.Sheets(i).Name
        if "add" in sheet_name or "13" in sheet_name:
            ws_src = wb_src.Sheets(i)
            print(f"Found target sheet: '{ws_src.Name}'", flush=True)
            break

    if ws_src is None:
        ws_src = wb_src.Sheets(2)
        print(f"Using 2nd sheet: '{ws_src.Name}'", flush=True)

    # Get used range
    used = ws_src.UsedRange
    last_row = used.Rows.Count
    last_col = used.Columns.Count
    print(f"Source data: {last_row} rows x {last_col} cols", flush=True)

    # Read all values
    data = used.Value
    print("Data read successfully.", flush=True)

    print("Opening destination xlsm...", flush=True)
    wb_dst = excel.Workbooks.Open(dst, UpdateLinks=0)
    ws_dst = wb_dst.Worksheets(1)
    print(f"Destination Sheet1 name: '{ws_dst.Name}'", flush=True)

    # Clear Sheet1
    ws_dst.Cells.Clear()
    print("Sheet1 cleared.", flush=True)

    # Paste data
    ws_dst.Range(ws_dst.Cells(1, 1), ws_dst.Cells(last_row, last_col)).Value = data
    print(f"Data pasted: {last_row} rows x {last_col} cols", flush=True)

    # Close source
    wb_src.Close(False)
    print("Source file closed.", flush=True)

    # Run macro
    print("Running macro...", flush=True)
    try:
        excel.Run("'메틀러토레도_CJ택배_매크로 (1).xlsm'!CJ_Macro.전체실행")
        macro_success = True
        print("Macro ran successfully (method 1).", flush=True)
    except Exception as e:
        macro_error_msg = str(e)
        print(f"Macro error method 1: {e}", flush=True)
        try:
            excel.Run("CJ_Macro.전체실행")
            macro_success = True
            print("Macro ran successfully (method 2).", flush=True)
        except Exception as e2:
            macro_error_msg += " | " + str(e2)
            print(f"Macro error method 2: {e2}", flush=True)

    # Save and close
    wb_dst.Save()
    wb_dst.Close(False)
    print("Destination saved and closed.", flush=True)

except Exception as e:
    print(f"FATAL ERROR: {e}", flush=True)
    import traceback
    traceback.print_exc()

finally:
    try:
        excel.DisplayAlerts = True
        excel.Quit()
        print("Excel quit.", flush=True)
    except:
        pass

print("\n--- Post-processing with openpyxl ---", flush=True)

try:
    wb_result = openpyxl.load_workbook(dst, read_only=True, keep_vba=True)
    print(f"Sheets: {wb_result.sheetnames}", flush=True)

    # Find CJ택배양식 sheet
    cj_sheet = None
    for name in wb_result.sheetnames:
        if "CJ" in name or "택배" in name or "양식" in name:
            cj_sheet = wb_result[name]
            print(f"Found result sheet: '{name}'", flush=True)
            break

    if cj_sheet is None:
        print("CJ택배양식 sheet not found. Available sheets:", wb_result.sheetnames, flush=True)
        cj_sheet = wb_result.active
        print(f"Using active sheet: '{cj_sheet.title}'", flush=True)

    rows = list(cj_sheet.iter_rows(values_only=True))
    # Find first non-empty row as header
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(cell is not None for cell in row):
            header_row = row
            header_idx = i
            break

    data_rows = [r for r in rows[header_idx+1:] if any(c is not None for c in r)]
    print(f"Total rows in result sheet (with header): {len(rows)}", flush=True)
    print(f"Data rows (non-empty): {len(data_rows)}", flush=True)

    if header_row:
        print(f"Header: {[str(h)[:15] if h else '' for h in header_row[:15]]}", flush=True)

        # Find key columns
        col_map = {}
        for j, h in enumerate(header_row):
            if h is None:
                continue
            h_str = str(h).strip()
            if h_str == "수령인":
                col_map["수령인"] = j
            elif "연락처" in h_str or ("수령인" in h_str and "연락처" in h_str):
                col_map["수령인연락처"] = j
            elif "주소" in h_str:
                col_map["수령인주소"] = j
            elif "DN" in h_str.upper():
                col_map["DN NO."] = j

        print(f"Key column positions: {col_map}", flush=True)

        print("\nFirst 5 data rows (key columns):", flush=True)
        for i, row in enumerate(data_rows[:5]):
            row_data = {}
            for col_name, col_idx in col_map.items():
                if col_idx < len(row):
                    row_data[col_name] = row[col_idx]
            print(f"  Row {i+1}: {row_data}", flush=True)

    wb_result.close()

except Exception as e:
    print(f"openpyxl error: {e}", flush=True)
    import traceback
    traceback.print_exc()

print("\n=== FINAL SUMMARY ===", flush=True)
print(f"Source rows read: {last_row} (including header)", flush=True)
print(f"Macro success: {macro_success}", flush=True)
if macro_error_msg:
    print(f"Macro error details: {macro_error_msg}", flush=True)
