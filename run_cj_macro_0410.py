import win32com.client as win32
import pythoncom
import openpyxl
import os

src = r"C:\Users\dkchoi\Downloads\0410 dn add .xlsx"
dst = r"C:\Users\dkchoi\Downloads\메틀러토레도_CJ택배_매크로 (1).xlsm"

excel = win32.Dispatch("Excel.Application")
excel.Visible = False
excel.DisplayAlerts = False

macro_success = False
macro_error_msg = ""

try:
    # Open source file
    wb_src = excel.Workbooks.Open(src)

    # Find "13시 add" sheet (should be 2nd sheet)
    ws_src = None
    for i in range(1, wb_src.Sheets.Count + 1):
        sheet_name = wb_src.Sheets(i).Name
        print(f"  Sheet {i}: {sheet_name}")
        if "add" in sheet_name:
            ws_src = wb_src.Sheets(i)
            print(f"Found sheet: {ws_src.Name}")
            break

    if ws_src is None:
        ws_src = wb_src.Sheets(2)
        print(f"Using 2nd sheet: {ws_src.Name}")

    # Get used range
    used = ws_src.UsedRange
    last_row = used.Rows.Count
    last_col = used.Columns.Count
    print(f"Source data: {last_row} rows x {last_col} cols")

    # Read all values
    data = used.Value

    # Open destination xlsm
    wb_dst = excel.Workbooks.Open(dst)
    ws_dst = wb_dst.Worksheets(1)  # Sheet1

    # Clear Sheet1
    ws_dst.Cells.Clear()

    # Paste data
    ws_dst.Range(ws_dst.Cells(1, 1), ws_dst.Cells(last_row, last_col)).Value = data
    print(f"Data pasted to Sheet1: {last_row} rows x {last_col} cols")

    # Close source
    wb_src.Close(False)

    # Run macro
    try:
        excel.Run("'메틀러토레도_CJ택배_매크로 (1).xlsm'!CJ_Macro.전체실행")
        macro_success = True
        print("Macro ran successfully")
    except Exception as e:
        macro_error_msg = str(e)
        print(f"Macro error (trying alternate call): {e}")
        try:
            excel.Run("CJ_Macro.전체실행")
            macro_success = True
            print("Macro ran (alternate)")
        except Exception as e2:
            macro_error_msg += " | " + str(e2)
            print(f"Macro alternate error: {e2}")

    wb_dst.Save()
    wb_dst.Close(False)
    print("DONE - file saved")

finally:
    excel.DisplayAlerts = True
    excel.Quit()

print("\n--- Post-processing with openpyxl ---")

# Read result with openpyxl
wb_result = openpyxl.load_workbook(dst, read_only=True, keep_vba=True)
print(f"Sheets in result: {wb_result.sheetnames}")

# Find CJ택배양식 sheet
cj_sheet = None
for name in wb_result.sheetnames:
    if "CJ" in name or "택배" in name or "양식" in name:
        cj_sheet = wb_result[name]
        print(f"Found result sheet: {name}")
        break

if cj_sheet is None:
    print("CJ택배양식 sheet not found, checking all sheets...")
    for name in wb_result.sheetnames:
        print(f"  - {name}")
    # Try first sheet as fallback
    cj_sheet = wb_result.active

# Count rows and show first 5
rows = list(cj_sheet.iter_rows(values_only=True))
# Find header row
header_row = None
header_idx = 0
for i, row in enumerate(rows):
    if row and any(cell is not None for cell in row):
        header_row = row
        header_idx = i
        break

print(f"\nTotal rows in sheet (including header): {len(rows)}")

# Find columns: 수령인, 수령인연락처, 수령인주소, DN NO.
if header_row:
    print(f"Header row (row {header_idx+1}): {[str(h)[:20] if h else '' for h in header_row[:10]]}")

    col_map = {}
    for j, h in enumerate(header_row):
        if h is None:
            continue
        h_str = str(h).strip()
        if "수령인" == h_str or "수령인" in h_str:
            if "수령인" not in col_map:
                col_map["수령인"] = j
        if "연락처" in h_str or "전화" in h_str:
            col_map["수령인연락처"] = j
        if "주소" in h_str:
            col_map["수령인주소"] = j
        if "DN" in h_str or "dn" in h_str.lower():
            col_map["DN NO."] = j

    print(f"Column map: {col_map}")

    # Show first 5 data rows
    print("\nFirst 5 data rows (key columns):")
    data_rows = [r for r in rows[header_idx+1:] if any(c is not None for c in r)]
    print(f"Data rows count: {len(data_rows)}")

    for i, row in enumerate(data_rows[:5]):
        row_data = {}
        for col_name, col_idx in col_map.items():
            if col_idx < len(row):
                row_data[col_name] = row[col_idx]
        print(f"  Row {i+1}: {row_data}")

wb_result.close()

print(f"\n=== SUMMARY ===")
print(f"Macro success: {macro_success}")
if macro_error_msg:
    print(f"Macro error: {macro_error_msg}")
