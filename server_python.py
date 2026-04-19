#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
os.environ.setdefault('PYTHONUTF8', '1')

import json, secrets, re, mimetypes, socket, io
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from datetime import datetime, timedelta
import urllib.request

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── 환경변수 ───────────────────────────────────────────────────────────
DATABASE_URL     = os.environ.get('DATABASE_URL', '')
STATIC           = os.path.join(os.path.dirname(__file__), 'public')
ADMIN_PW         = os.environ.get('ADMIN_PW', 'admin1234')
STAFF_PW         = os.environ.get('STAFF_PW', 'staff1234')
COMPANY          = os.environ.get('COMPANY', '출고 인수증명 시스템')
PORT             = int(os.environ.get('PORT', 3000))
_tg_raw          = os.environ.get('TELEGRAM_TOKEN', '')
_tg_match        = re.search(r'(\d+:[A-Za-z0-9_-]+)', _tg_raw)
TELEGRAM_TOKEN   = _tg_match.group(1) if _tg_match else ''
TELEGRAM_CHAT_ID = re.sub(r'[^0-9-]', '', os.environ.get('TELEGRAM_CHAT_ID', ''))

valid_tokens       = set()
valid_staff_tokens = set()

def send_telegram(msg):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID: return
    try:
        data = json.dumps({'chat_id': TELEGRAM_CHAT_ID, 'text': msg, 'parse_mode': 'HTML'}).encode()
        req  = urllib.request.Request(
            f'https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage',
            data=data, headers={'Content-Type': 'application/json'})
        urllib.request.urlopen(req, timeout=5)
    except Exception as e:
        print(f'Telegram 알림 오류: {e}')

# ── DB 레이어 ─────────────────────────────────────────────────────────
if DATABASE_URL:
    import psycopg2, psycopg2.extras

    def _conn():
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor, sslmode='require')

    def init_db():
        c = _conn(); cur = c.cursor()
        # 기존 테이블
        cur.execute('''CREATE TABLE IF NOT EXISTS delivery_records (
            id SERIAL PRIMARY KEY, order_no TEXT UNIQUE NOT NULL,
            delivery_date TEXT, arrival_time TEXT, product_name TEXT,
            quantity TEXT, customer_company TEXT, customer_address TEXT,
            receiver_name TEXT, receiver_phone TEXT, driver_name TEXT,
            driver_phone TEXT, vehicle_no TEXT, wait_time TEXT, work_time TEXT,
            waste_collection TEXT, extra_locations TEXT, notes TEXT,
            driver_signature TEXT, receiver_signature TEXT, signed_at TEXT,
            status TEXT DEFAULT 'draft',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        for col in ['work_fee TEXT','return_fee TEXT','delivery_note TEXT','vehicle_type TEXT',
                    'origin TEXT','origin_address TEXT','contact_person TEXT','contact_phone TEXT',
                    'transport_type TEXT','dest_sido TEXT','dest_sigun TEXT','origin_sido TEXT','origin_sigun TEXT']:
            try: cur.execute(f"ALTER TABLE delivery_records ADD COLUMN IF NOT EXISTS {col}")
            except: pass
        cur.execute('''CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)''')
        # ── 신규 테이블 ──
        cur.execute('''CREATE TABLE IF NOT EXISTS todos (
            id SERIAL PRIMARY KEY, title TEXT NOT NULL,
            done INTEGER DEFAULT 0, priority TEXT DEFAULT 'normal',
            due_date TEXT, notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS dispatch_items (
            id SERIAL PRIMARY KEY, client TEXT, item_date TEXT,
            origin TEXT, destination TEXT, cargo TEXT, vehicle_type TEXT,
            carrier_fee TEXT, carrier_work TEXT, shipper_fee TEXT, shipper_work TEXT,
            assigned_to TEXT, contact TEXT, carrier_name TEXT, notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS claims (
            id SERIAL PRIMARY KEY, dn_no TEXT, client TEXT,
            claim_type TEXT, description TEXT, status TEXT DEFAULT 'open',
            resolution TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS calendar_events (
            id SERIAL PRIMARY KEY, title TEXT NOT NULL,
            event_date TEXT, event_time TEXT, event_type TEXT DEFAULT 'general',
            notes TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS memos (
            id SERIAL PRIMARY KEY, title TEXT, content TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS vendors (
            id SERIAL PRIMARY KEY, name TEXT NOT NULL, vendor_type TEXT,
            contact TEXT, phone TEXT, address TEXT, notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS shipper_reqs (
            id SERIAL PRIMARY KEY, client TEXT, requirement TEXT,
            priority TEXT DEFAULT 'normal', status TEXT DEFAULT 'active',
            notes TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS daily_journals (
            id SERIAL PRIMARY KEY, client TEXT NOT NULL,
            journal_date TEXT NOT NULL, data_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(client, journal_date))''')
        cur.execute('''CREATE TABLE IF NOT EXISTS bonded_records (
            id SERIAL PRIMARY KEY, record_date TEXT, record_type TEXT,
            bl_no TEXT, mrn TEXT, customer TEXT, product TEXT,
            quantity TEXT, weight TEXT, notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS ot_records (
            id SERIAL PRIMARY KEY, work_date TEXT NOT NULL,
            start_time TEXT, end_time TEXT, work_content TEXT,
            ot_hours REAL DEFAULT 0, meal_ticket TEXT DEFAULT 'X',
            notes TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        # OT 시드 데이터 (테이블이 비어있을 때만 삽입)
        cur.execute('SELECT COUNT(*) as cnt FROM ot_records')
        row = cur.fetchone(); cnt = dict(row).get('cnt', 0) if row else 0
        if cnt == 0:
            ot_seed = [
                ('2026-03-25','18:00','20:10','메틀러토레도 입고', 2.2,'X'),
                ('2026-03-27','18:00','18:50','메틀러토레도 출고', 0.8,'X'),
                ('2026-03-31','18:00','18:50','메틀러토레도 입고', 0.8,'X'),
                ('2026-04-01','18:00','21:50','메틀러토레도 출고', 3.8,'O'),
                ('2026-04-13','18:00','18:50','캐논메디칼 출고',   0.8,'X'),
                ('2026-04-14','22:00','22:40','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-15','18:30','19:00','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-18','13:20','14:50','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-19','07:30','09:00','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-19','11:30','15:40','보세근무',           4.2,'O'),
            ]
            for d in ot_seed:
                cur.execute(
                    'INSERT INTO ot_records (work_date,start_time,end_time,work_content,ot_hours,meal_ticket) VALUES (%s,%s,%s,%s,%s,%s)', d)
        c.commit(); c.close()

    def db_fetch(sql, params=()):
        sql = sql.replace('?','%s')
        c = _conn(); cur = c.cursor(); cur.execute(sql, params)
        row = cur.fetchone(); c.close()
        return dict(row) if row else None

    def db_fetchall(sql, params=()):
        sql = sql.replace('?','%s')
        c = _conn(); cur = c.cursor(); cur.execute(sql, params)
        rows = cur.fetchall(); c.close()
        return [dict(r) for r in rows]

    def db_exec(sql, params=(), returning=False):
        sql = sql.replace('?','%s')
        c = _conn(); cur = c.cursor(); cur.execute(sql, params)
        result = None
        if returning:
            row = cur.fetchone()
            result = dict(row)['id'] if row else None
        c.commit(); c.close()
        return result

    def db_insert(sql, params=()):
        sql = sql.replace('?','%s')
        if 'RETURNING' not in sql.upper(): sql = sql.strip() + ' RETURNING id'
        c = _conn(); cur = c.cursor(); cur.execute(sql, params)
        row = cur.fetchone(); c.commit(); c.close()
        return dict(row)['id'] if row else None

    INTEGRITY_EXC = psycopg2.IntegrityError

else:
    import sqlite3
    DB_PATH = os.path.join(os.path.dirname(__file__), 'delivery.db')

    def _conn():
        c = sqlite3.connect(DB_PATH); c.row_factory = sqlite3.Row; return c

    def init_db():
        c = _conn()
        c.execute('''CREATE TABLE IF NOT EXISTS delivery_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT, order_no TEXT UNIQUE NOT NULL,
            delivery_date TEXT, arrival_time TEXT, product_name TEXT,
            quantity TEXT, customer_company TEXT, customer_address TEXT,
            receiver_name TEXT, receiver_phone TEXT, driver_name TEXT,
            driver_phone TEXT, vehicle_no TEXT, wait_time TEXT, work_time TEXT,
            waste_collection TEXT, extra_locations TEXT, notes TEXT,
            driver_signature TEXT, receiver_signature TEXT, signed_at TEXT,
            status TEXT DEFAULT 'draft',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        for col in ['work_fee TEXT','return_fee TEXT','delivery_note TEXT','vehicle_type TEXT',
                    'origin TEXT','origin_address TEXT','contact_person TEXT','contact_phone TEXT',
                    'transport_type TEXT','dest_sido TEXT','dest_sigun TEXT','origin_sido TEXT','origin_sigun TEXT']:
            try: c.execute(f'ALTER TABLE delivery_records ADD COLUMN {col}'); c.commit()
            except: pass
        c.execute('''CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)''')
        # ── 신규 테이블 ──
        c.execute('''CREATE TABLE IF NOT EXISTS todos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL,
            done INTEGER DEFAULT 0, priority TEXT DEFAULT 'normal',
            due_date TEXT, notes TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS dispatch_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, client TEXT, item_date TEXT,
            origin TEXT, destination TEXT, cargo TEXT, vehicle_type TEXT,
            carrier_fee TEXT, carrier_work TEXT, shipper_fee TEXT, shipper_work TEXT,
            assigned_to TEXT, contact TEXT, carrier_name TEXT, notes TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS claims (
            id INTEGER PRIMARY KEY AUTOINCREMENT, dn_no TEXT, client TEXT,
            claim_type TEXT, description TEXT, status TEXT DEFAULT 'open',
            resolution TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS calendar_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL,
            event_date TEXT, event_time TEXT, event_type TEXT DEFAULT 'general',
            notes TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS memos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT, content TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS vendors (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
            vendor_type TEXT, contact TEXT, phone TEXT, address TEXT, notes TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS shipper_reqs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, client TEXT, requirement TEXT,
            priority TEXT DEFAULT 'normal', status TEXT DEFAULT 'active',
            notes TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS daily_journals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, client TEXT NOT NULL,
            journal_date TEXT NOT NULL, data_json TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(client, journal_date))''')
        c.execute('''CREATE TABLE IF NOT EXISTS bonded_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT, record_date TEXT, record_type TEXT,
            bl_no TEXT, mrn TEXT, customer TEXT, product TEXT,
            quantity TEXT, weight TEXT, notes TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        c.execute('''CREATE TABLE IF NOT EXISTS ot_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT, work_date TEXT NOT NULL,
            start_time TEXT, end_time TEXT, work_content TEXT,
            ot_hours REAL DEFAULT 0, meal_ticket TEXT DEFAULT 'X',
            notes TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP)''')
        # OT 시드 데이터 (테이블이 비어있을 때만 삽입)
        cnt = c.execute('SELECT COUNT(*) FROM ot_records').fetchone()[0]
        if cnt == 0:
            ot_seed = [
                ('2026-03-25','18:00','20:10','메틀러토레도 입고', 2.2,'X'),
                ('2026-03-27','18:00','18:50','메틀러토레도 출고', 0.8,'X'),
                ('2026-03-31','18:00','18:50','메틀러토레도 입고', 0.8,'X'),
                ('2026-04-01','18:00','21:50','메틀러토레도 출고', 3.8,'O'),
                ('2026-04-13','18:00','18:50','캐논메디칼 출고',   0.8,'X'),
                ('2026-04-14','22:00','22:40','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-15','18:30','19:00','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-18','13:20','14:50','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-19','07:30','09:00','캐논메디칼 긴급출고',3.0,'X'),
                ('2026-04-19','11:30','15:40','보세근무',           4.2,'O'),
            ]
            c.executemany(
                'INSERT INTO ot_records (work_date,start_time,end_time,work_content,ot_hours,meal_ticket) VALUES (?,?,?,?,?,?)', ot_seed)
        c.commit(); c.close()

    def db_fetch(sql, params=()):
        c = _conn(); row = c.execute(sql, params).fetchone(); c.close()
        return dict(row) if row else None

    def db_fetchall(sql, params=()):
        c = _conn(); rows = c.execute(sql, params).fetchall(); c.close()
        return [dict(r) for r in rows]

    def db_exec(sql, params=(), returning=False):
        c = _conn(); cur = c.execute(sql, params)
        result = cur.lastrowid if returning else None
        c.commit(); c.close()
        return result

    def db_insert(sql, params=()):
        c = _conn(); cur = c.execute(sql, params)
        new_id = cur.lastrowid; c.commit(); c.close()
        return new_id

    INTEGRITY_EXC = sqlite3.IntegrityError


# ── Excel 생성 ─────────────────────────────────────────────────────────
def _border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def make_mettler_excel(data, journal_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '메틀러 업무일지'

    # 열 너비
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10

    hdr_fill = _fill('BDD7EE')
    title_font = Font(name='맑은 고딕', bold=True, size=16)
    bold = Font(name='맑은 고딕', bold=True, size=10)
    normal = Font(name='맑은 고딕', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 행 높이
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 8
    for r in range(3, 17): ws.row_dimensions[r].height = 22

    # ── 제목 ──
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = '메틀러토레도 일일 업무일지'
    c.font = title_font; c.alignment = center; c.fill = _fill('D9E1F2')
    for col in ['F','G','H','I']:
        ws[f'{col}1'].fill = _fill('F2F2F2')
    ws['F1'].value = '결재'; ws['F1'].font = bold; ws['F1'].alignment = center
    ws['G1'].value = '차  장'; ws['G1'].font = bold; ws['G1'].alignment = center
    ws['H1'].value = '이  사'; ws['H1'].font = bold; ws['H1'].alignment = center
    ws['I1'].value = '대표이사'; ws['I1'].font = bold; ws['I1'].alignment = center

    # 날짜 파싱
    try:
        dt = datetime.strptime(journal_date, '%Y-%m-%d')
        wd = ['월','화','수','목','금','토','일'][dt.weekday()]
        date_str = f"{dt.year}년 {dt.month}월 {dt.day}일 {wd}요일"
    except:
        date_str = journal_date

    # ── 기본 정보 ──
    def set_label(cell, text):
        ws[cell].value = text; ws[cell].font = bold
        ws[cell].alignment = center; ws[cell].fill = hdr_fill
        ws[cell].border = _border()

    def set_val(cell_range_or_cell, text):
        cell = ws[cell_range_or_cell] if ':' not in cell_range_or_cell else None
        if cell:
            cell.value = text; cell.font = normal
            cell.alignment = center; cell.border = _border()

    def merge_val(rng, text):
        ws.merge_cells(rng)
        c = ws[rng.split(':')[0]]
        c.value = text; c.font = normal; c.alignment = center; c.border = _border()

    set_label('A3', '작성일자'); merge_val('B3:E3', date_str)
    set_label('F3', '업무시작'); merge_val('G3:I3', data.get('work_start','08:30'))

    set_label('A4', '작성자'); merge_val('B4:E4', data.get('writer',''))
    set_label('F4', '업무종료'); merge_val('G4:I4', data.get('work_end','18:00'))

    set_label('A5', '인원')
    ws.merge_cells('B5:C5'); ws['B5'].value = f"{data.get('staff_count','')}명"
    ws['B5'].font = normal; ws['B5'].alignment = center; ws['B5'].border = _border()
    set_label('F5', '비고'); merge_val('G5:I5', data.get('header_notes',''))

    # ── 입고 ──
    for col, txt in [('A6','구분'),('B6','신한입고'),('C6','내국입고'),('D6','특수입고'),('E6','비고')]:
        ws[col].value = txt; ws[col].font = bold; ws[col].alignment = center
        ws[col].fill = hdr_fill; ws[col].border = _border()
    ws.merge_cells('F6:I6'); ws['F6'].value = ''; ws['F6'].border = _border()

    imp = data.get('import_plan', {})
    imd = data.get('import_done', {})
    for row, key, label in [(7, 'import_plan', '입고예정'), (8, 'import_done', '입고완료')]:
        d = data.get(key, {})
        set_label(f'A{row}', label)
        for col, k in [('B', 'sinhan'), ('C', 'domestic'), ('D', 'special')]:
            ws[f'{col}{row}'].value = d.get(k, 0)
            ws[f'{col}{row}'].font = normal; ws[f'{col}{row}'].alignment = center
            ws[f'{col}{row}'].border = _border()
        merge_val(f'E{row}:I{row}', d.get('notes',''))

    # ── 출고 ──
    for col, txt in [('A9','구분'),('B9','택배출고 건수'),('C9','차량출고 건수'),('D9','픽업출고 건수'),('E9','비고')]:
        ws[col].value = txt; ws[col].font = bold; ws[col].alignment = center
        ws[col].fill = hdr_fill; ws[col].border = _border()
    ws.merge_cells('F9:I9'); ws['F9'].border = _border()

    exp = data.get('export', {})
    set_label('A10', '출고')
    for col, k in [('B','courier'), ('C','vehicle'), ('D','pickup')]:
        ws[f'{col}10'].value = exp.get(k, 0)
        ws[f'{col}10'].font = normal; ws[f'{col}10'].alignment = center
        ws[f'{col}10'].border = _border()
    merge_val('E10:I10', exp.get('notes',''))

    # ── 특이사항 ──
    ws.merge_cells('A11:I11')
    ws['A11'].value = '특이사항'; ws['A11'].font = bold; ws['A11'].alignment = center
    ws['A11'].fill = hdr_fill; ws['A11'].border = _border()

    ws.row_dimensions[12].height = 60
    ws.merge_cells('A12:I12')
    ws['A12'].value = data.get('special_notes', '')
    ws['A12'].font = normal; ws['A12'].alignment = left; ws['A12'].border = _border()

    # ── 로고 영역 ──
    ws.row_dimensions[14].height = 28
    ws.merge_cells('A14:C14')
    ws['A14'].value = '[ KOREA AEO ]'; ws['A14'].font = Font(name='맑은 고딕', bold=True, size=10, color='003366')
    ws['A14'].alignment = center
    ws.merge_cells('D14:I14')
    ws['D14'].value = 'ACT LOGISTICS Co., Ltd.'; ws['D14'].font = Font(name='맑은 고딕', bold=True, size=11, color='003366')
    ws['D14'].alignment = left

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def make_chanel_excel(data, journal_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '샤넬 업무일지'

    # 열 너비
    cols_w = {'A':8,'B':8,'C':8,'D':10,'E':10,'F':10,'G':10,'H':10,
              'I':3,'J':10,'K':10,'L':10,'M':10,'N':10}
    for col, w in cols_w.items():
        ws.column_dimensions[col].width = w

    hdr_fill  = _fill('BDD7EE')
    grn_fill  = _fill('E2EFDA')
    pnk_fill  = _fill('FCE4D6')
    title_font = Font(name='맑은 고딕', bold=True, size=15)
    bold  = Font(name='맑은 고딕', bold=True, size=9)
    norm  = Font(name='맑은 고딕', size=9)
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for r in [1,2]: ws.row_dimensions[r].height = 30
    for r in range(3,30): ws.row_dimensions[r].height = 18

    try:
        dt = datetime.strptime(journal_date, '%Y-%m-%d')
        wd = ['월','화','수','목','금','토','일'][dt.weekday()]
        date_str = f"{dt.year}년 {dt.month}월 {dt.day}일 {wd}요일"
    except:
        date_str = journal_date

    def bc(cell, val, fnt=None, aln=None, fil=None, brd=True):
        c = ws[cell]
        c.value = val
        c.font = fnt or norm; c.alignment = aln or center
        if fil: c.fill = fil
        if brd: c.border = _border()

    def mc(rng, val, fnt=None, aln=None, fil=None):
        ws.merge_cells(rng)
        bc(rng.split(':')[0], val, fnt, aln, fil)

    # ── 제목 ──
    mc('A1:H1', '샤넬 업무일지', title_font, center, _fill('D9E1F2'))
    for col in ['I','J','K','L','M','N']:
        ws[f'{col}1'].fill = _fill('F2F2F2'); ws[f'{col}1'].border = _border()
    ws['J1'].value = '결재'; ws['J1'].font = bold; ws['J1'].alignment = center
    ws['K1'].value = '차  장'; ws['K1'].font = bold; ws['K1'].alignment = center
    ws['L1'].value = '이  사'; ws['L1'].font = bold; ws['L1'].alignment = center
    ws['M1'].value = '대표이사'; ws['M1'].font = bold; ws['M1'].alignment = center

    mc('A2:C2', '작성일자', bold, center, hdr_fill)
    mc('D2:H2', date_str, norm, center)
    mc('I2:J2', '업무시작', bold, center, hdr_fill)
    mc('K2:N2', data.get('work_start','08:30'), norm, center)

    mc('A3:C3', '작성자', bold, center, hdr_fill)
    mc('D3:H3', data.get('writer',''), norm, center)
    mc('I3:J3', '업무종료', bold, center, hdr_fill)
    mc('K3:N3', data.get('work_end','18:00'), norm, center)

    # ── 섹션 헤더 ──
    mc('A4:H4', '패션', bold, center, grn_fill)
    mc('I4:N4', '화장품', bold, center, pnk_fill)

    # ── 항목 헤더 ──
    headers_l = [('A5',''),('B5','구분'),('C5',''),('D5','내국'),('E5','보세'),('F5',''),('G5',''),('H5','')]
    for cell, val in headers_l:
        bc(cell, val, bold, center, hdr_fill)

    # 패션 입/출/재고
    fa = data.get('fashion', {})
    co = data.get('cosmetics', {})

    rows_def = [
        (6,  '입고', '내국(BOX)', fa.get('import_domestic_box',0), '내국(PLT)', fa.get('import_domestic_plt',0)),
        (7,  '',     '보세(BOX)', fa.get('import_bonded_box',0),   '보세(PLT)', fa.get('import_bonded_plt',0)),
        (8,  '출고', '차량(건)', fa.get('export_vehicle_count',0), '차량(BOX)', fa.get('export_vehicle_qty',0)),
        (9,  '',     '택배(건)', fa.get('export_courier_count',0), '택배(BOX)', fa.get('export_courier_qty',0)),
        (10, '현재고','BOX',     fa.get('stock_box',0),            'PLT',       fa.get('stock_plt',0)),
    ]
    for row, cat, lbl1, val1, lbl2, val2 in rows_def:
        bc(f'A{row}', cat, bold, center, hdr_fill if cat else None)
        bc(f'B{row}', lbl1, norm, center)
        bc(f'C{row}', val1, norm, center)
        bc(f'D{row}', lbl2, norm, center)
        bc(f'E{row}', val2, norm, center)
        ws.merge_cells(f'F{row}:H{row}'); ws[f'F{row}'].border = _border()

    # 화장품 입/출/재고 (오른쪽)
    co_rows = [
        (6,  '입고', '내국(PCS)', co.get('import_domestic_pcs',0), '내국(PLT)', co.get('import_domestic_plt',0)),
        (7,  '',     '보세(PCS)', co.get('import_bonded_pcs',0),   '보세(PLT)', co.get('import_bonded_plt',0)),
        (8,  '출고', '차량(건)',  co.get('export_vehicle_count',0), '차량(PCS)', co.get('export_vehicle_qty',0)),
        (9,  '',     '택배(건)',  co.get('export_courier_count',0), '택배(PCS)', co.get('export_courier_qty',0)),
        (10, '현재고','PCS',      co.get('stock_pcs',0),            'PLT',       co.get('stock_plt',0)),
    ]
    for row, cat, lbl1, val1, lbl2, val2 in co_rows:
        bc(f'I{row}',  cat, bold, center, hdr_fill if cat else None)
        bc(f'J{row}',  lbl1, norm, center)
        bc(f'K{row}',  val1, norm, center)
        bc(f'L{row}',  lbl2, norm, center)
        bc(f'M{row}',  val2, norm, center)
        ws[f'N{row}'].border = _border()

    # ── 어제 재고 ──
    bc('A11', '어제재고', bold, center, hdr_fill)
    mc('B11:C11', fa.get('yesterday_stock',''), norm, center)
    bc('D11', 'PLT', bold, center, hdr_fill)
    bc('E11', fa.get('yesterday_plt',''), norm, center)
    ws.merge_cells('F11:H11'); ws['F11'].border = _border()
    bc('I11', '어제재고', bold, center, hdr_fill)
    mc('J11:K11', co.get('yesterday_stock',''), norm, center)
    bc('L11', 'PLT', bold, center, hdr_fill)
    bc('M11', co.get('yesterday_plt',''), norm, center)
    ws['N11'].border = _border()

    # ── 기타 요약 (우측) ──
    extra = data.get('extra', {})
    mc('A12:H12', '특이사항', bold, center, hdr_fill)
    mc('I12:N12', '기타 현황', bold, center, hdr_fill)
    ws.row_dimensions[13].height = 70
    ws.merge_cells('A13:H13')
    ws['A13'].value = data.get('special_notes','')
    ws['A13'].font = norm; ws['A13'].alignment = left; ws['A13'].border = _border()

    extra_items = [
        ('전산등록 X', extra.get('register_x','')),
        ('중복', extra.get('duplicate','')),
        ('WMS 수량', extra.get('wms_qty','')),
        ('재고이동', extra.get('stock_move','')),
        ('패션+화장품 PLT', extra.get('total_plt','')),
    ]
    for i, (lbl, val) in enumerate(extra_items):
        row = 13 + i
        if row == 13:
            bc('I13', lbl, bold, center, hdr_fill)
            mc('J13:N13', val, norm, center)
        else:
            ws.row_dimensions[row].height = 18
            mc(f'I{row}:J{row}', lbl, bold, center, hdr_fill)
            mc(f'K{row}:N{row}', val, norm, center)

    # ── 로고 ──
    logo_row = 20
    ws.row_dimensions[logo_row].height = 28
    mc(f'A{logo_row}:C{logo_row}', '[ KOREA AEO ]',
       Font(name='맑은 고딕', bold=True, size=10, color='003366'), center)
    mc(f'D{logo_row}:N{logo_row}', 'ACT LOGISTICS Co., Ltd.',
       Font(name='맑은 고딕', bold=True, size=11, color='003366'), left)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ── 핸들러 ────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, data, filename, mime='application/octet-stream'):
        self.send_response(200)
        self.send_header('Content-Type', mime)
        self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.send_header('Content-Length', len(data))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(data)

    def read_body(self):
        length = int(self.headers.get('Content-Length', 0))
        return json.loads(self.rfile.read(length)) if length else {}

    def get_token(self):
        return self.headers.get('Authorization','').replace('Bearer ','').strip()

    def token_ok(self, admin_only=False):
        t = self.get_token()
        if t and t in valid_tokens: return True
        if not admin_only and t and t in valid_staff_tokens: return True
        return False

    def _serve_static(self, req_path):
        safe = req_path.lstrip('/')
        if not safe: safe = 'index.html'
        file_path = os.path.normpath(os.path.join(STATIC, safe))
        if not file_path.startswith(os.path.normpath(STATIC)):
            self.send_response(403); self.end_headers(); return
        if os.path.isdir(file_path):
            file_path = os.path.join(file_path, 'index.html')
        if not os.path.isfile(file_path):
            self.send_response(404); self.send_header('Content-Type','text/plain')
            self.end_headers(); self.wfile.write(b'404 Not Found'); return
        mime, _ = mimetypes.guess_type(file_path)
        if not mime: mime = 'application/octet-stream'
        with open(file_path, 'rb') as f: data = f.read()
        self.send_response(200)
        self.send_header('Content-Type', mime)
        self.send_header('Content-Length', len(data))
        self.end_headers()
        self.wfile.write(data)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,PATCH,DELETE,OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        self.end_headers()

    # ── GET ──────────────────────────────────────────────────────────
    def do_GET(self):
      try:
        p   = urlparse(self.path)
        path = p.path.rstrip('/')
        qs  = parse_qs(p.query)
        g   = lambda k, d='': (qs.get(k, [''])[0] or d)

        if path == '/api/config':
            return self.send_json({'companyName': COMPANY})

        if path == '/api/auth/check':
            if self.token_ok(): return self.send_json({'ok': True})
            return self.send_json({'error':'Unauthorized'}, 401)

        if path == '/api/test-telegram':
            if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
                return self.send_json({'error':'환경변수 없음'})
            try:
                data = json.dumps({'chat_id': TELEGRAM_CHAT_ID, 'text': '🔔 Render 서버 테스트!'}).encode()
                req  = urllib.request.Request(
                    f'https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage',
                    data=data, headers={'Content-Type': 'application/json'})
                res  = urllib.request.urlopen(req, timeout=10)
                return self.send_json({'ok': True, 'result': json.loads(res.read().decode())})
            except Exception as e:
                return self.send_json({'error': str(e)})

        # 기사 검색 (공개)
        if path == '/api/sign/search':
            order_no = g('order_no').strip()
            if not order_no: return self.send_json({'data': []})
            row = db_fetch(
                'SELECT id,order_no,delivery_date,product_name,quantity,customer_company,'
                'customer_address,receiver_name,driver_name,driver_phone,vehicle_no,'
                'status,extra_locations,wait_time,work_time,work_fee,notes '
                'FROM delivery_records WHERE order_no=? OR order_no LIKE ? OR order_no LIKE ? OR order_no LIKE ?',
                (order_no, f'{order_no} 외%', f'{order_no}외%', f'{order_no}%'))
            return self.send_json({'data': [row] if row else []})

        # 운송사 검색 (공개)
        if path == '/api/carrier/search':
            order_no = g('order_no').strip()
            if not order_no: return self.send_json({'data': []})
            row = db_fetch(
                'SELECT id,order_no,delivery_date,product_name,quantity,customer_company,'
                'customer_address,receiver_name,wait_time,work_fee,extra_locations,waste_collection '
                'FROM delivery_records WHERE order_no=? OR order_no LIKE ? OR order_no LIKE ? OR order_no LIKE ?',
                (order_no, f'{order_no} 외%', f'{order_no}외%', f'{order_no}%'))
            return self.send_json({'data': [row] if row else []})

        # ── 대시보드 통계 (admin) ──
        if path == '/api/stats':
            if not self.token_ok(admin_only=True):
                return self.send_json({'error':'Unauthorized'}, 401)
            today = (datetime.utcnow() + timedelta(hours=9)).strftime('%Y-%m-%d')
            total_today    = (db_fetch('SELECT COUNT(*) as c FROM delivery_records WHERE delivery_date=?', (today,)) or {}).get('c',0)
            signed_today   = (db_fetch('SELECT COUNT(*) as c FROM delivery_records WHERE delivery_date=? AND status=?', (today,'signed')) or {}).get('c',0)
            pending_today  = (db_fetch('SELECT COUNT(*) as c FROM delivery_records WHERE delivery_date=? AND status=?', (today,'draft')) or {}).get('c',0)
            open_claims    = (db_fetch("SELECT COUNT(*) as c FROM claims WHERE status='open'") or {}).get('c',0)
            pending_todos  = (db_fetch("SELECT COUNT(*) as c FROM todos WHERE done=0") or {}).get('c',0)
            bonded_in      = (db_fetch("SELECT COUNT(*) as c FROM bonded_records WHERE record_date=? AND record_type='반입'", (today,)) or {}).get('c',0)
            bonded_out     = (db_fetch("SELECT COUNT(*) as c FROM bonded_records WHERE record_date=? AND record_type='반출'", (today,)) or {}).get('c',0)
            recent_signed  = db_fetchall(
                "SELECT order_no,customer_company,signed_at FROM delivery_records WHERE status='signed' ORDER BY signed_at DESC LIMIT 5")
            return self.send_json({
                'today': today, 'total_today': total_today,
                'signed_today': signed_today, 'pending_today': pending_today,
                'open_claims': open_claims, 'pending_todos': pending_todos,
                'bonded_in': bonded_in, 'bonded_out': bonded_out,
                'recent_signed': recent_signed
            })

        # ── 오더 목록 (admin/staff) ──
        if path == '/api/records':
            if not self.token_ok(): return self.send_json({'error':'Unauthorized'}, 401)
            search = g('search'); dn = g('dn'); company = g('company')
            date_from = g('dateFrom'); date_to = g('dateTo'); status = g('status')
            limit = int(g('limit','1000') or 1000)
            sql = 'SELECT * FROM delivery_records WHERE 1=1'; params = []
            if search:
                sql += ' AND (order_no LIKE ? OR customer_company LIKE ? OR receiver_name LIKE ? OR product_name LIKE ?)'
                s = f'%{search}%'; params += [s,s,s,s]
            if dn:        sql += ' AND order_no LIKE ?';         params.append(f'%{dn}%')
            if company:   sql += ' AND customer_company LIKE ?'; params.append(f'%{company}%')
            if date_from: sql += ' AND delivery_date >= ?';      params.append(date_from)
            if date_to:   sql += ' AND delivery_date <= ?';      params.append(date_to)
            if status:    sql += ' AND status=?';                params.append(status)
            sql += ' ORDER BY created_at DESC LIMIT ?'; params.append(limit)
            return self.send_json({'data': db_fetchall(sql, params)})

        # 오더 단건
        m = re.match(r'^/api/records/(\d+)$', path)
        if m:
            if not self.token_ok(): return self.send_json({'error':'Unauthorized'}, 401)
            row = db_fetch('SELECT * FROM delivery_records WHERE id=?', (m.group(1),))
            if not row: return self.send_json({'error':'Not found'}, 404)
            return self.send_json(row)

        # ── 일일 업무일지 ──
        if path == '/api/journals':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            client = g('client'); date = g('date')
            sql = 'SELECT * FROM daily_journals WHERE 1=1'; params=[]
            if client: sql += ' AND client=?'; params.append(client)
            if date:   sql += ' AND journal_date=?'; params.append(date)
            sql += ' ORDER BY journal_date DESC LIMIT 100'
            rows = db_fetchall(sql, params)
            for r in rows:
                if r.get('data_json'):
                    try: r['data'] = json.loads(r['data_json'])
                    except: r['data'] = {}
            return self.send_json({'data': rows})

        # 업무일지 Excel 다운로드
        if path == '/api/journals/excel':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            if not EXCEL_OK: return self.send_json({'error':'openpyxl 미설치'}, 500)
            client = g('client'); date = g('date')
            row = db_fetch('SELECT * FROM daily_journals WHERE client=? AND journal_date=?', (client, date))
            if not row: return self.send_json({'error':'데이터 없음'}, 404)
            try: d = json.loads(row['data_json'])
            except: d = {}
            if client == 'mettler':
                xls = make_mettler_excel(d, date)
                fname = f'메틀러토레도_업무일지_{date}.xlsx'
            else:
                xls = make_chanel_excel(d, date)
                fname = f'샤넬코리아_업무일지_{date}.xlsx'
            return self.send_file(xls, fname,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # ── 보세 반입/출 ──
        if path == '/api/bonded':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            date = g('date'); rtype = g('type')
            sql = 'SELECT * FROM bonded_records WHERE 1=1'; params=[]
            if date:  sql += ' AND record_date=?';  params.append(date)
            if rtype: sql += ' AND record_type=?'; params.append(rtype)
            sql += ' ORDER BY created_at DESC'
            return self.send_json({'data': db_fetchall(sql, params)})

        # ── OT 내역 ──
        if path == '/api/ot':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            # 3~4월 기간 OT 시드 데이터 없으면 자동 삽입
            try:
                seed_chk = db_fetchall("SELECT COUNT(*) as cnt FROM ot_records WHERE work_date >= ? AND work_date <= ?", ('2026-03-01','2026-04-30'))
                cnt_val = int(seed_chk[0].get('cnt', 0)) if seed_chk else 0
            except: cnt_val = 0
            if cnt_val == 0:
                ot_seed = [
                    ('2026-03-25','18:00','20:10','메틀러토레도 입고', 2.2,'X'),
                    ('2026-03-27','18:00','18:50','메틀러토레도 출고', 0.8,'X'),
                    ('2026-03-31','18:00','18:50','메틀러토레도 입고', 0.8,'X'),
                    ('2026-04-01','18:00','21:50','메틀러토레도 출고', 3.8,'O'),
                    ('2026-04-13','18:00','18:50','캐논메디칼 출고',   0.8,'X'),
                    ('2026-04-14','22:00','22:40','캐논메디칼 긴급출고',3.0,'X'),
                    ('2026-04-15','18:30','19:00','캐논메디칼 긴급출고',3.0,'X'),
                    ('2026-04-18','13:20','14:50','캐논메디칼 긴급출고',3.0,'X'),
                    ('2026-04-19','07:30','09:00','캐논메디칼 긴급출고',3.0,'X'),
                    ('2026-04-19','11:30','15:40','보세근무',           4.2,'O'),
                ]
                for d in ot_seed:
                    db_insert(
                        'INSERT INTO ot_records (work_date,start_time,end_time,work_content,ot_hours,meal_ticket) VALUES (?,?,?,?,?,?)', d)
            ym   = g('ym')    # YYYY-MM 월별 필터
            frm  = g('from')  # YYYY-MM-DD 시작일
            to   = g('to')    # YYYY-MM-DD 종료일
            sql = 'SELECT * FROM ot_records WHERE 1=1'; params=[]
            if ym:  sql += ' AND work_date LIKE ?'; params.append(ym + '%')
            if frm: sql += ' AND work_date >= ?';  params.append(frm)
            if to:  sql += ' AND work_date <= ?';  params.append(to)
            sql += ' ORDER BY work_date ASC, id ASC'
            rows = db_fetchall(sql, params)
            total_hours = sum(float(r.get('ot_hours') or 0) for r in rows)
            total_hours = round(total_hours, 1)
            meal_count  = sum(1 for r in rows if (r.get('meal_ticket') or 'X') == 'O')
            return self.send_json({'data': rows, 'total_hours': total_hours, 'meal_count': meal_count})

        # ── 제네릭 목록 API (admin) ──
        RESOURCE_MAP = {
            'todos':     ('todos',          'created_at DESC'),
            'dispatch':  ('dispatch_items', 'item_date DESC, created_at DESC'),
            'claims':    ('claims',         'created_at DESC'),
            'calendar':  ('calendar_events','event_date ASC'),
            'memos':     ('memos',          'updated_at DESC'),
            'vendors':   ('vendors',        'name ASC'),
            'shipper-reqs': ('shipper_reqs','created_at DESC'),
        }
        m2 = re.match(r'^/api/(todos|dispatch|claims|calendar|memos|vendors|shipper-reqs)$', path)
        if m2:
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            resource = m2.group(1)
            table, order = RESOURCE_MAP[resource]
            sql = f'SELECT * FROM {table} WHERE 1=1'
            params = []
            client = g('client'); status = g('status'); month = g('month')
            if client and resource in ('dispatch','claims','shipper-reqs'):
                sql += ' AND client=?'; params.append(client)
            if status and resource == 'claims':
                sql += ' AND status=?'; params.append(status)
            if month and resource == 'calendar':
                sql += ' AND event_date LIKE ?'; params.append(f'{month}%')
            sql += f' ORDER BY {order}'
            return self.send_json({'data': db_fetchall(sql, params)})

        # 단건 조회
        m3 = re.match(r'^/api/(todos|dispatch|claims|calendar|memos|vendors|shipper-reqs|bonded)/(\d+)$', path)
        if m3:
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            resource, rid = m3.group(1), m3.group(2)
            TABLES = {'todos':'todos','dispatch':'dispatch_items','claims':'claims',
                      'calendar':'calendar_events','memos':'memos','vendors':'vendors',
                      'shipper-reqs':'shipper_reqs','bonded':'bonded_records'}
            row = db_fetch(f"SELECT * FROM {TABLES[resource]} WHERE id=?", (rid,))
            if not row: return self.send_json({'error':'Not found'}, 404)
            return self.send_json(row)

        self._serve_static(p.path)
      except Exception as e:
        try: self.send_response(500); self.end_headers()
        except: pass

    # ── POST ─────────────────────────────────────────────────────────
    def do_POST(self):
      try:
        p    = urlparse(self.path).path.rstrip('/')
        body = self.read_body()

        # 인증
        if p == '/api/auth':
            pw = body.get('password','')
            db_admin = (db_fetch('SELECT value FROM app_settings WHERE key=?', ('admin_pw',)) or {}).get('value') or ADMIN_PW
            db_staff = (db_fetch('SELECT value FROM app_settings WHERE key=?', ('staff_pw',)) or {}).get('value') or STAFF_PW
            if pw == db_admin:
                t = secrets.token_hex(32); valid_tokens.add(t)
                return self.send_json({'token': t, 'role': 'admin'})
            elif pw == db_staff:
                t = secrets.token_hex(32); valid_staff_tokens.add(t)
                return self.send_json({'token': t, 'role': 'staff'})
            return self.send_json({'error':'비밀번호가 틀렸습니다.'}, 401)

        # 기사 서명 (공개)
        m = re.match(r'^/api/sign/(\d+)$', p)
        if m:
            rid = m.group(1)
            rec = db_fetch('SELECT id,status FROM delivery_records WHERE id=?', (rid,))
            if not rec: return self.send_json({'error':'오더를 찾을 수 없습니다.'}, 404)
            if rec['status'] == 'signed': return self.send_json({'error':'이미 서명 완료된 오더입니다.'}, 400)
            if not body.get('driver_signature') or not body.get('receiver_signature'):
                return self.send_json({'error':'서명 데이터가 없습니다.'}, 400)
            now = (datetime.utcnow() + timedelta(hours=9)).strftime('%Y-%m-%d %H:%M:%S')
            db_exec('''UPDATE delivery_records SET
                driver_name=?,driver_phone=?,vehicle_no=?,receiver_name=?,
                driver_signature=?,receiver_signature=?,signed_at=?,status='signed'
                WHERE id=?''',
                (body.get('driver_name'), body.get('driver_phone'), body.get('vehicle_no'),
                 body.get('receiver_name'), body['driver_signature'], body['receiver_signature'], now, rid))
            rec2 = db_fetch('SELECT * FROM delivery_records WHERE id=?', (rid,))
            if rec2:
                send_telegram(
                    f'📦 <b>서명 완료 알림</b>\n'
                    f'DN번호: {rec2.get("order_no","")}\n'
                    f'고객사: {rec2.get("customer_company","")}\n'
                    f'제품명: {rec2.get("product_name","")}\n'
                    f'수량: {rec2.get("quantity","")}\n'
                    f'수취인: {rec2.get("receiver_name","")}\n'
                    f'기사: {body.get("driver_name","")}\n'
                    f'완료시각: {now}')
            return self.send_json({'success': True})

        # 오더 등록 (admin/staff)
        if p == '/api/records':
            if not self.token_ok(): return self.send_json({'error':'Unauthorized'}, 401)
            try:
                new_id = db_insert(
                    '''INSERT INTO delivery_records
                        (order_no,delivery_date,arrival_time,product_name,quantity,
                         customer_company,customer_address,receiver_name,receiver_phone,
                         driver_name,driver_phone,vehicle_no,wait_time,work_time,
                         waste_collection,extra_locations,notes,
                         delivery_note,vehicle_type,
                         origin,origin_address,contact_person,contact_phone)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (body['order_no'], body.get('delivery_date'), body.get('arrival_time'),
                     body.get('product_name'), body.get('quantity'), body.get('customer_company'),
                     body.get('customer_address'), body.get('receiver_name'), body.get('receiver_phone'),
                     body.get('driver_name'), body.get('driver_phone'), body.get('vehicle_no'),
                     body.get('wait_time'), body.get('work_time'), body.get('waste_collection'),
                     body.get('extra_locations'), body.get('notes'),
                     body.get('delivery_note'), body.get('vehicle_type'),
                     body.get('origin'), body.get('origin_address'),
                     body.get('contact_person'), body.get('contact_phone')))
                return self.send_json({'id': new_id, **body})
            except INTEGRITY_EXC:
                return self.send_json({'error': f"DN번호 [{body.get('order_no')}]이 이미 등록되어 있습니다."}, 400)

        # 일일 업무일지 저장 (admin)
        if p == '/api/journals':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            client = body.get('client'); date = body.get('journal_date')
            data_json = json.dumps(body.get('data', {}), ensure_ascii=False)
            existing = db_fetch('SELECT id FROM daily_journals WHERE client=? AND journal_date=?', (client, date))
            if existing:
                db_exec('UPDATE daily_journals SET data_json=? WHERE client=? AND journal_date=?',
                        (data_json, client, date))
                return self.send_json({'success': True, 'action': 'updated'})
            else:
                new_id = db_insert('INSERT INTO daily_journals (client, journal_date, data_json) VALUES (?,?,?)',
                                   (client, date, data_json))
                return self.send_json({'success': True, 'id': new_id, 'action': 'created'})

        # 보세 반입/출 등록 (admin)
        if p == '/api/bonded':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            new_id = db_insert(
                'INSERT INTO bonded_records (record_date,record_type,bl_no,mrn,customer,product,quantity,weight,notes) VALUES (?,?,?,?,?,?,?,?,?)',
                (body.get('record_date'), body.get('record_type'), body.get('bl_no'),
                 body.get('mrn'), body.get('customer'), body.get('product'),
                 body.get('quantity'), body.get('weight'), body.get('notes')))
            return self.send_json({'success': True, 'id': new_id})

        # ── OT POST ──
        if p == '/api/ot':
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            new_id = db_insert(
                'INSERT INTO ot_records (work_date,start_time,end_time,work_content,ot_hours,meal_ticket,notes) VALUES (?,?,?,?,?,?,?)',
                (body.get('work_date'), body.get('start_time'), body.get('end_time'),
                 body.get('work_content'), body.get('ot_hours', 0),
                 body.get('meal_ticket', 'X'), body.get('notes', '')))
            return self.send_json({'success': True, 'id': new_id})

        # ── 제네릭 POST (admin) ──
        INSERT_MAP = {
            '/api/todos':        ('todos',          ['title','done','priority','due_date','notes']),
            '/api/dispatch':     ('dispatch_items', ['client','item_date','origin','destination','cargo','vehicle_type','carrier_fee','carrier_work','shipper_fee','shipper_work','assigned_to','contact','carrier_name','notes']),
            '/api/claims':       ('claims',         ['dn_no','client','claim_type','description','status','resolution']),
            '/api/calendar':     ('calendar_events',['title','event_date','event_time','event_type','notes']),
            '/api/memos':        ('memos',          ['title','content']),
            '/api/vendors':      ('vendors',        ['name','vendor_type','contact','phone','address','notes']),
            '/api/shipper-reqs': ('shipper_reqs',   ['client','requirement','priority','status','notes']),
        }
        if p in INSERT_MAP:
            if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
            table, fields = INSERT_MAP[p]
            cols   = [f for f in fields if f in body]
            if not cols: return self.send_json({'error':'데이터 없음'}, 400)
            vals   = [body[c] for c in cols]
            sql    = f"INSERT INTO {table} ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})"
            new_id = db_insert(sql, vals)
            # 긴급 클레임 텔레그램 알림
            if p == '/api/claims' and body.get('claim_type') == '긴급':
                send_telegram(
                    f'🚨 <b>긴급 클레임 등록</b>\n'
                    f'DN번호: {body.get("dn_no","")}\n'
                    f'화주: {body.get("client","")}\n'
                    f'내용: {body.get("description","")}')
            return self.send_json({'success': True, 'id': new_id})

        self.send_json({'error':'Not found'}, 404)
      except Exception as e:
        try: self.send_json({'error': str(e)}, 500)
        except: pass

    # ── PATCH ────────────────────────────────────────────────────────
    def do_PATCH(self):
      try:
        p = urlparse(self.path).path.rstrip('/')

        # 운송사 작업 (공개)
        mc = re.match(r'^/api/carrier/(\d+)$', p)
        if mc:
            body    = self.read_body()
            allowed = ['wait_time','work_fee','extra_locations','waste_collection']
            data    = {k:v for k,v in body.items() if k in allowed}
            if not data: return self.send_json({'error':'변경할 항목 없음'}, 400)
            sets = ', '.join(f'{k}=?' for k in data.keys())
            vals = list(data.values()) + [mc.group(1)]
            db_exec(f'UPDATE delivery_records SET {sets} WHERE id=?', vals)
            return self.send_json({'success': True})

        if not self.token_ok(): return self.send_json({'error':'Unauthorized'}, 401)
        body = self.read_body()

        # 비밀번호 변경
        if p == '/api/settings/password':
            if not self.token_ok(admin_only=True):
                return self.send_json({'error':'관리자만 변경할 수 있습니다.'}, 403)
            if body.get('admin_pw'):
                db_exec('INSERT INTO app_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',
                        ('admin_pw', body['admin_pw']))
            if body.get('staff_pw'):
                db_exec('INSERT INTO app_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',
                        ('staff_pw', body['staff_pw']))
            return self.send_json({'success': True})

        # 오더 수정
        m = re.match(r'^/api/records/(\d+)$', p)
        if m:
            sets = ', '.join(f'{k}=?' for k in body.keys())
            vals = list(body.values()) + [m.group(1)]
            db_exec(f'UPDATE delivery_records SET {sets} WHERE id=?', vals)
            return self.send_json({'success': True})

        # ── 제네릭 PATCH (admin) ──
        if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)
        PATCH_MAP = {
            'todos':'todos', 'dispatch':'dispatch_items', 'claims':'claims',
            'calendar':'calendar_events', 'memos':'memos', 'vendors':'vendors',
            'shipper-reqs':'shipper_reqs', 'bonded':'bonded_records', 'ot':'ot_records'
        }
        m2 = re.match(r'^/api/(todos|dispatch|claims|calendar|memos|vendors|shipper-reqs|bonded|ot)/(\d+)$', p)
        if m2:
            resource, rid = m2.group(1), m2.group(2)
            table = PATCH_MAP[resource]
            # memos: updated_at 자동 갱신
            if resource == 'memos':
                body['updated_at'] = (datetime.utcnow() + timedelta(hours=9)).strftime('%Y-%m-%d %H:%M:%S')
            sets = ', '.join(f'{k}=?' for k in body.keys())
            vals = list(body.values()) + [rid]
            db_exec(f'UPDATE {table} SET {sets} WHERE id=?', vals)
            return self.send_json({'success': True})

        self.send_json({'error':'Not found'}, 404)
      except Exception as e:
        try: self.send_response(500); self.end_headers()
        except: pass

    # ── DELETE ───────────────────────────────────────────────────────
    def do_DELETE(self):
      try:
        p = urlparse(self.path).path.rstrip('/')
        if not self.token_ok(admin_only=True): return self.send_json({'error':'Unauthorized'}, 401)

        m = re.match(r'^/api/records/(\d+)$', p)
        if m:
            db_exec('DELETE FROM delivery_records WHERE id=?', (m.group(1),))
            return self.send_json({'success': True})

        DELETE_MAP = {
            'todos':'todos', 'dispatch':'dispatch_items', 'claims':'claims',
            'calendar':'calendar_events', 'memos':'memos', 'vendors':'vendors',
            'shipper-reqs':'shipper_reqs', 'bonded':'bonded_records', 'ot':'ot_records'
        }
        m2 = re.match(r'^/api/(todos|dispatch|claims|calendar|memos|vendors|shipper-reqs|bonded|ot)/(\d+)$', p)
        if m2:
            resource, rid = m2.group(1), m2.group(2)
            db_exec(f"DELETE FROM {DELETE_MAP[resource]} WHERE id=?", (rid,))
            return self.send_json({'success': True})

        self.send_json({'error':'Not found'}, 404)
      except Exception as e:
        try: self.send_response(500); self.end_headers()
        except: pass


# ── 실행 ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    local_ip = '127.0.0.1'
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80)); local_ip = s.getsockname()[0]; s.close()
    except: pass

    server = ThreadingHTTPServer(('0.0.0.0', PORT), Handler)
    server.allow_reuse_address = True
    sep = '=' * 46
    print(sep)
    print('  [토탈 물류 프로그램] 가동 중')
    print(sep)
    print(f'  PC  : http://localhost:{PORT}')
    print(f'  LAN : http://{local_ip}:{PORT}')
    print(f'  DB  : {"PostgreSQL (Cloud)" if DATABASE_URL else "SQLite (Local)"}')
    print(f'  Excel: {"openpyxl OK" if EXCEL_OK else "openpyxl 미설치 - pip install openpyxl"}')
    print('  Ctrl+C 로 종료')
    print(sep)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('Server stopped.')
